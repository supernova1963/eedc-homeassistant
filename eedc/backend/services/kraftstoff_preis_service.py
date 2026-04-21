"""
Kraftstoffpreis-Service — EU Weekly Oil Bulletin.

Lädt wöchentliche nationale Durchschnittspreise für Euro-Super 95
von der EU-Kommission (XLSX) und speichert sie in TagesZusammenfassung.

Quelle: https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en
Daten seit 2005, wöchentlich (Montag), alle EU-Länder.
Preise: Euro-Super 95 inkl. Steuern, €/1000L → umgerechnet in €/L.

CH ist nicht in der EU — für CH-Anlagen wird AT als Approximation verwendet.
"""

import io
import logging
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Optional

import httpx

logger = logging.getLogger(__name__)

# Stabile URL zur History-XLSX (aktualisiert wöchentlich, feste UUID)
OIL_BULLETIN_URL = (
    "https://energy.ec.europa.eu/document/download/"
    "906e60ca-8b6a-44e7-8589-652854d2fd3f_en"
    "?filename=Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx"
)

# EEDC standort_land → Oil Bulletin Länderkürzel
LAND_MAPPING = {
    "DE": "DE",
    "AT": "AT",
    "IT": "IT",
    "CH": "AT",    # CH nicht in EU → AT als Näherung
    "FR": "FR",
    "NL": "NL",
    "BE": "BE",
    "LU": "LU",
    "ES": "ES",
    "PT": "PT",
    "GR": "GR",
    "PL": "PL",
    "CZ": "CZ",
    "DK": "DK",
    "SE": "SE",
    "FI": "FI",
    "IE": "IE",
    "HU": "HU",
    "RO": "RO",
    "BG": "BG",
    "HR": "HR",
    "SK": "SK",
    "SI": "SI",
    "LT": "LT",
    "LV": "LV",
    "EE": "EE",
    "CY": "CY",
    "MT": "MT",
}

# Cache: {land: [(datum, preis_euro_l), ...]} — wird beim ersten Abruf befüllt
_cache: dict[str, list[tuple[date, float]]] = {}
_cache_timestamp: Optional[datetime] = None
CACHE_TTL_HOURS = 24  # Wöchentliche Daten → täglicher Refresh reicht


async def _download_xlsx() -> Optional[bytes]:
    """Lädt die Oil Bulletin XLSX herunter."""
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.get(OIL_BULLETIN_URL, follow_redirects=True)
            resp.raise_for_status()
            data = resp.content
            if len(data) < 10000:
                logger.warning("Oil Bulletin XLSX zu klein (%d bytes), vermutlich Fehler", len(data))
                return None
            logger.info("Oil Bulletin XLSX heruntergeladen: %.1f MB", len(data) / 1024 / 1024)
            return data
    except Exception as e:
        logger.error("Oil Bulletin Download fehlgeschlagen: %s", e)
        return None


def _parse_xlsx(data: bytes) -> dict[str, list[tuple[date, float]]]:
    """
    Parst die Oil Bulletin XLSX und extrahiert Euro-Super 95 Preise pro Land.

    Returns:
        {land_code: [(datum, preis_euro_liter), ...]} sortiert nach Datum absteigend
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl nicht installiert — pip install openpyxl")
        return {}

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Prices with taxes"]

    # Header-Zeile 0: Spaltenbezeichnungen parsen
    # Format: "{LAND}_price_with_tax_euro95" → Spaltenindex merken
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    land_spalten: dict[str, int] = {}
    for i, val in enumerate(header):
        if val and isinstance(val, str) and "price_with_tax_euro95" in val:
            land_code = val.split("_")[0]
            if land_code not in ("EU", "EUR"):  # EU-Durchschnitt überspringen
                land_spalten[land_code] = i

    if not land_spalten:
        logger.error("Keine Euro-Super-95-Spalten in Oil Bulletin gefunden")
        wb.close()
        return {}

    logger.info("Oil Bulletin: %d Länder gefunden (%s)",
                len(land_spalten), ", ".join(sorted(land_spalten.keys())))

    # Datenzeilen ab Zeile 4 (0-indexed: Row 3)
    ergebnis: dict[str, list[tuple[date, float]]] = defaultdict(list)

    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = list(row)
        datum_raw = vals[0]
        if not datum_raw or not hasattr(datum_raw, "year"):
            continue

        datum = datum_raw.date() if isinstance(datum_raw, datetime) else datum_raw

        for land, col_idx in land_spalten.items():
            if col_idx < len(vals):
                preis_1000l = vals[col_idx]
                if preis_1000l is not None and isinstance(preis_1000l, (int, float)) and preis_1000l > 100:
                    # €/1000L → €/L
                    preis_l = round(preis_1000l / 1000, 3)
                    ergebnis[land].append((datum, preis_l))

    wb.close()

    # Sortiere nach Datum absteigend (neueste zuerst)
    for land in ergebnis:
        ergebnis[land].sort(key=lambda x: x[0], reverse=True)

    total = sum(len(v) for v in ergebnis.values())
    logger.info("Oil Bulletin geparst: %d Preise für %d Länder", total, len(ergebnis))

    return dict(ergebnis)


async def get_kraftstoffpreise(land: str = "DE", force_refresh: bool = False) -> list[tuple[date, float]]:
    """
    Gibt wöchentliche Kraftstoffpreise für ein Land zurück.

    Args:
        land: EEDC standort_land Code (DE, AT, CH, IT, ...)
        force_refresh: Cache ignorieren und neu laden

    Returns:
        [(datum, preis_euro_liter), ...] sortiert nach Datum absteigend.
        Leere Liste wenn Daten nicht verfügbar.
    """
    global _cache, _cache_timestamp

    oil_land = LAND_MAPPING.get(land, land)

    if not force_refresh and _cache and _cache_timestamp:
        alter_h = (datetime.now() - _cache_timestamp).total_seconds() / 3600
        if alter_h < CACHE_TTL_HOURS and oil_land in _cache:
            return _cache[oil_land]

    # Download + Parse
    xlsx_data = await _download_xlsx()
    if not xlsx_data:
        # Fallback auf Cache wenn vorhanden
        return _cache.get(oil_land, [])

    _cache = _parse_xlsx(xlsx_data)
    _cache_timestamp = datetime.now()

    return _cache.get(oil_land, [])


def get_preis_fuer_datum(preise: list[tuple[date, float]], ziel_datum: date) -> Optional[float]:
    """
    Findet den passenden Wochenpreis für ein gegebenes Datum.

    Oil Bulletin erscheint montags — der Preis gilt für die Woche davor.
    Wir nehmen den nächsten Preis der <= ziel_datum ist.
    """
    for datum, preis in preise:
        if datum <= ziel_datum:
            return preis
    # Kein passender Preis (Datum liegt vor 2005)
    return None


async def backfill_kraftstoffpreise(
    anlage_id: int,
    land: str,
    db,  # AsyncSession
    von: Optional[date] = None,
    bis: Optional[date] = None,
) -> dict:
    """
    Befüllt kraftstoffpreis_euro in TagesZusammenfassung für eine Anlage.

    Lädt Oil Bulletin Preise und schreibt den passenden Wochenpreis
    in alle TagesZusammenfassung-Zeilen die noch keinen Preis haben.

    Returns:
        {"aktualisiert": int, "land": str, "zeitraum": str}
    """
    from sqlalchemy import select, and_
    from sqlalchemy.orm.attributes import flag_modified
    from backend.models.tages_energie_profil import TagesZusammenfassung

    preise = await get_kraftstoffpreise(land)
    if not preise:
        return {"aktualisiert": 0, "fehler": "Keine Kraftstoffpreise verfügbar"}

    oil_land = LAND_MAPPING.get(land, land)

    # Alle TagesZusammenfassungen ohne Kraftstoffpreis laden
    query = select(TagesZusammenfassung).where(
        TagesZusammenfassung.anlage_id == anlage_id,
        TagesZusammenfassung.kraftstoffpreis_euro.is_(None),
    )
    if von:
        query = query.where(TagesZusammenfassung.datum >= von)
    if bis:
        query = query.where(TagesZusammenfassung.datum <= bis)

    result = await db.execute(query)
    rows = result.scalars().all()

    if not rows:
        return {"aktualisiert": 0, "land": oil_land, "hinweis": "Alle Tage haben bereits einen Preis"}

    aktualisiert = 0
    for tz in rows:
        preis = get_preis_fuer_datum(preise, tz.datum)
        if preis is not None:
            tz.kraftstoffpreis_euro = preis
            aktualisiert += 1

    if aktualisiert > 0:
        await db.commit()
        logger.info("Kraftstoffpreis-Backfill: %d Tage für Anlage %d (%s) aktualisiert",
                     aktualisiert, anlage_id, oil_land)

    return {
        "aktualisiert": aktualisiert,
        "land": oil_land,
        "zeitraum": f"{rows[0].datum} – {rows[-1].datum}" if rows else "",
    }
